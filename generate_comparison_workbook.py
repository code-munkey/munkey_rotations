#!/usr/bin/env python3
"""
Generate Excel workbook for rotation comparison across LLMs.

Creates a multi-sheet Excel file with one sheet per class.
Each sheet contains comparison rows for all specs and sources.

Usage:
    python generate_comparison_workbook.py
    python generate_comparison_workbook.py --output rotation_comparison.xlsx
"""

import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Will generate CSV files instead.")
    print("Install with: pip install openpyxl")


# All WoW classes and their specs
CLASSES_AND_SPECS = {
    "Death Knight": ["Blood", "Frost", "Unholy"],
    "Demon Hunter": ["Havoc", "Vengeance"],
    "Druid": ["Balance", "Feral", "Guardian", "Restoration"],
    "Evoker": ["Devastation", "Preservation", "Augmentation"],
    "Hunter": ["Beast Mastery", "Marksmanship", "Survival"],
    "Mage": ["Arcane", "Fire", "Frost"],
    "Monk": ["Brewmaster", "Mistweaver", "Windwalker"],
    "Paladin": ["Holy", "Protection", "Retribution"],
    "Priest": ["Discipline", "Holy", "Shadow"],
    "Rogue": ["Assassination", "Outlaw", "Subtlety"],
    "Shaman": ["Elemental", "Enhancement", "Restoration"],
    "Warlock": ["Affliction", "Demonology", "Destruction"],
    "Warrior": ["Arms", "Fury", "Protection"],
}

SOURCES = ["Wowhead", "Icy Veins", "Method.gg"]
ROTATION_TYPES = ["ST", "AOE"]
LLMS = ["Kimi (This LLM)", "Claude Opus", "Gemini 3 Pro"]


def create_excel_workbook(output_path: Path) -> None:
    """Create multi-sheet Excel workbook."""
    if not HAS_OPENPYXL:
        print("Error: openpyxl required for Excel output")
        return
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet, we'll create our own
    wb.remove(wb.active)
    
    # Create summary sheet first
    summary = wb.create_sheet("Summary", 0)
    create_summary_sheet(summary)
    
    # Create a sheet for each class
    for class_name, specs in CLASSES_AND_SPECS.items():
        sheet = wb.create_sheet(class_name)
        create_class_sheet(sheet, class_name, specs)
    
    wb.save(output_path)
    print(f"Excel workbook saved to: {output_path}")


def create_summary_sheet(sheet):
    """Create the summary/overview sheet."""
    sheet.title = "Summary"
    
    # Title
    sheet['A1'] = "Rotation Comparison - Summary"
    sheet['A1'].font = Font(size=16, bold=True)
    
    # Description
    sheet['A3'] = "LLMs Being Compared:"
    for i, llm in enumerate(LLMS, start=4):
        sheet[f'A{i}'] = f"  - {llm}"
    
    sheet['A8'] = "Sources Being Compared:"
    for i, source in enumerate(SOURCES, start=9):
        sheet[f'A{i}'] = f"  - {source}"
    
    sheet['A13'] = "Rotation Types:"
    sheet['A14'] = "  - ST (Single Target)"
    sheet['A15'] = "  - AOE (Area of Effect)"
    
    # Scoring guide
    sheet['A17'] = "Scoring Guide (1-5 or Pass/Fail):"
    sheet['A17'].font = Font(bold=True)
    
    scoring = [
        ("5 / Pass", "Perfect - Follows all patterns, proper syntax, handles edge cases"),
        ("4", "Good - Minor issues, mostly correct"),
        ("3", "Acceptable - Works but has notable issues"),
        ("2", "Poor - Significant problems, may not function correctly"),
        ("1 / Fail", "Broken - Syntax errors, wrong patterns, won't work"),
    ]
    
    for i, (score, desc) in enumerate(scoring, start=18):
        sheet[f'A{i}'] = score
        sheet[f'B{i}'] = desc
    
    # Common failure points
    sheet['A24'] = "Common Failure Points to Check:"
    sheet['A24'].font = Font(bold=True)
    
    failures = [
        "Uses '==' instead of '=' for equality",
        "movement_allowed inside variables section (should be root level)",
        "Calling lists/main explicitly (main is auto-executed)",
        "Missing spell morphs (override) - CHECK morph_database/",
        "Wrong empowered spell handling (missing ignore_usable/casting_check)",
        "Incorrect hero talent detection",
        "Missing parentheses for & | precedence",
        "Wrong config/variable reference (config.X vs var.X)",
    ]
    
    for i, failure in enumerate(failures, start=25):
        sheet[f'A{i}'] = f"  - {failure}"
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 70


def create_class_sheet(sheet, class_name: str, specs: list):
    """Create a sheet for a specific class."""
    
    # Header row
    headers = ["Spec", "Source", "Type"] + LLMS + ["Notes"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    row = 2
    for spec in specs:
        for source in SOURCES:
            for rot_type in ROTATION_TYPES:
                sheet.cell(row=row, column=1, value=spec)
                sheet.cell(row=row, column=2, value=source)
                sheet.cell(row=row, column=3, value=rot_type)
                
                # Empty cells for LLM scores (to be filled manually or via evaluation)
                for col in range(4, 4 + len(LLMS)):
                    sheet.cell(row=row, column=col, value="")
                
                # Notes column
                sheet.cell(row=row, column=4 + len(LLMS), value="")
                
                row += 1
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for r in range(1, row):
        for c in range(1, len(headers) + 1):
            sheet.cell(row=r, column=c).border = thin_border
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 20  # Spec
    sheet.column_dimensions['B'].width = 15  # Source
    sheet.column_dimensions['C'].width = 10  # Type
    for i, llm in enumerate(LLMS, start=4):
        sheet.column_dimensions[get_column_letter(i)].width = 18
    sheet.column_dimensions[get_column_letter(len(headers))].width = 40  # Notes


def create_csv_files(output_dir: Path) -> None:
    """Create separate CSV files for each class (fallback if openpyxl not available)."""
    output_dir.mkdir(exist_ok=True)
    
    for class_name, specs in CLASSES_AND_SPECS.items():
        filename = output_dir / f"{class_name.lower().replace(' ', '_')}_comparison.csv"
        
        lines = []
        lines.append("Spec,Source,Type,Kimi (This LLM),Claude Opus,Gemini 3 Pro,Notes")
        
        for spec in specs:
            for source in SOURCES:
                for rot_type in ROTATION_TYPES:
                    lines.append(f"{spec},{source},{rot_type},,,,")
        
        filename.write_text('\n'.join(lines))
        print(f"Created: {filename}")
    
    # Create summary CSV
    summary_file = output_dir / "_summary.csv"
    summary_lines = [
        "Rotation Comparison Summary",
        "",
        "LLMs Being Compared:",
    ]
    for llm in LLMS:
        summary_lines.append(f"  - {llm}")
    
    summary_lines.extend([
        "",
        "Sources Being Compared:",
    ])
    for source in SOURCES:
        summary_lines.append(f"  - {source}")
    
    summary_lines.extend([
        "",
        "Scoring Guide (1-5 or Pass/Fail):",
        "5/Pass,Perfect - Follows all patterns proper syntax handles edge cases",
        "4,Good - Minor issues mostly correct",
        "3,Acceptable - Works but has notable issues",
        "2,Poor - Significant problems may not function correctly",
        "1/Fail,Broken - Syntax errors wrong patterns won't work",
        "",
        "Common Failure Points:",
        "Uses == instead of = for equality",
        "movement_allowed inside variables section",
        "Calling lists/main explicitly",
        "Missing spell morphs (override)",
        "Wrong empowered spell handling",
        "Incorrect hero talent detection",
        "Missing parentheses for & | precedence",
    ])
    
    summary_file.write_text('\n'.join(summary_lines))
    print(f"Created: {summary_file}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate rotation comparison workbook"
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=Path("rotation_comparison.xlsx"),
        help="Output file path (default: rotation_comparison.xlsx)"
    )
    parser.add_argument(
        "--format", "-f",
        choices=["excel", "csv"],
        default="excel" if HAS_OPENPYXL else "csv",
        help="Output format"
    )
    
    args = parser.parse_args()
    
    if args.format == "excel" and HAS_OPENPYXL:
        create_excel_workbook(args.output)
    else:
        if args.format == "excel" and not HAS_OPENPYXL:
            print("openpyxl not available, falling back to CSV format")
        output_dir = args.output.parent / (args.output.stem + "_csv")
        create_csv_files(output_dir)


if __name__ == "__main__":
    main()
